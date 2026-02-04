"""
European Political Donations Tracker
Search for company donations across multiple European jurisdictions
"""

import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import re
from typing import Optional
import xml.etree.ElementTree as ET

# Page config
st.set_page_config(
    page_title="European Political Donations Tracker",
    page_icon="💰",
    layout="wide"
)

st.title("💰 European Political Donations Tracker")
st.markdown("""
Search for company or individual donations to political parties across European jurisdictions.
Enter a company name to search across available databases.
""")

# Data source info
DATA_SOURCES = {
    "uk": {
        "name": "UK Electoral Commission",
        "coverage": "Last 5 years (API has data from 2001)",
        "threshold": "£11,180 (central), £2,230 (accounting units)",
        "url": "https://search.electoralcommission.org.uk",
        "includes": "Parties AND individual MPs/Lords"
    },
    "germany": {
        "name": "German Bundestag",
        "coverage": "2002-present (immediate disclosure >€35K)",
        "threshold": "€35,000 (immediate), €10,000 (annual reports)",
        "url": "https://www.bundestag.de/parlament/parteienfinanzierung",
        "includes": "Parties only (MPs cannot accept monetary donations)"
    },
    "austria": {
        "name": "Austrian Court of Audit (Rechnungshof)",
        "coverage": "2019-present",
        "threshold": "€500 (disclosure), €2,500 (immediate)",
        "url": "https://www.rechnungshof.gv.at",
        "includes": "Parties only"
    },
    "italy": {
        "name": "Italian Parliament / Transparency International",
        "coverage": "2018-present",
        "threshold": "€500 (disclosure), €5,000 (donor name required)",
        "url": "https://soldiepolitica.it",
        "includes": "Parties only (MP data fragmented/separate)"
    },
    "netherlands": {
        "name": "Dutch Ministry of Interior (BZK)",
        "coverage": "2023-present",
        "threshold": "€10,000 (immediate disclosure), €1,000 (annual)",
        "url": "https://www.rijksoverheid.nl/onderwerpen/democratie/rol-politieke-partijen/giften-en-subsidies-politieke-partijen",
        "includes": "Parties only (candidate data separate)"
    },
    "estonia": {
        "name": "ERJK (Political Parties Financing Surveillance Committee)",
        "coverage": "2013-present",
        "threshold": "All donations disclosed (no minimum)",
        "url": "https://www.erjk.ee/en",
        "includes": "Parties only (corporate donations banned)",
        "note": "Only natural persons can donate"
    },
    "latvia": {
        "name": "KNAB (Corruption Prevention and Combating Bureau)",
        "coverage": "2015-present",
        "threshold": "All donations disclosed",
        "url": "https://info.knab.gov.lv/lv/db/ziedojumi/",
        "includes": "Parties only (corporate donations banned since 2004)",
        "note": "Only natural persons can donate - real-time search"
    },
    "lithuania": {
        "name": "VRK (Central Electoral Commission)",
        "coverage": "2013-present",
        "threshold": "All donations disclosed",
        "url": "https://data.gov.lt/datasets/2016/",
        "includes": "Campaign donations only (corporate donations banned)",
        "note": "Only citizens can donate"
    },
    "eu": {
        "name": "EU Authority for Political Parties",
        "coverage": "2018-present",
        "threshold": "€12,000",
        "url": "https://www.appf.europa.eu",
        "includes": "EU-level parties only",
        "mep_resources": {
            "gifts_register": "https://www.europarl.europa.eu/meps/en/about/meps",
            "integrity_watch": "https://www.integritywatch.eu/mepincomes",
            "note": "MEPs declare gifts >€150 and outside income separately"
        }
    }
}

# Sidebar with data source info
with st.sidebar:
    st.header("📊 Data Sources")
    for key, source in DATA_SOURCES.items():
        with st.expander(f"🔹 {source['name']}"):
            st.write(f"**Coverage:** {source['coverage']}")
            st.write(f"**Threshold:** {source['threshold']}")
            st.write(f"**Includes:** {source.get('includes', 'Parties')}")
            st.link_button("Visit source", source['url'])
    
    st.divider()
    st.markdown("### ⚙️ Search Settings")
    search_years = st.slider(
        "Years to search",
        min_value=1,
        max_value=10,
        value=5,
        help="Number of years back to search"
    )
    
    st.divider()
    st.markdown("""
    ### About
    This tool searches publicly available political donation 
    records across European jurisdictions.
    
    **Note:** Different countries have different disclosure 
    thresholds and reporting requirements. Results may not 
    be comprehensive.
    """)


# ============= UK ELECTORAL COMMISSION =============

def search_uk_donations(query: str, years: int = 5) -> pd.DataFrame:
    """
    Search UK Electoral Commission donations database.
    API endpoint allows CSV export with query parameters.
    Searches both Political Parties and Regulated Donees (MPs, Lords, etc.)
    Supports Boolean queries (AND, OR, NOT).
    """
    base_url = "https://search.electoralcommission.org.uk/api/csv/Donations"
    
    # Calculate date range
    end_date = datetime.now()
    start_date = end_date - timedelta(days=years * 365)
    
    # For Boolean queries, we need to search for each term separately and combine
    if is_boolean_query(query):
        search_terms = get_search_terms(query)
        all_dfs = []
        
        for term in search_terms:
            params = {
                "start": 0,
                "rows": 500,
                "query": term,
                "sort": "AcceptedDate",
                "order": "desc",
                "date": "Accepted",
                "from": start_date.strftime("%Y-%m-%d"),
                "to": end_date.strftime("%Y-%m-%d"),
                "prePoll": "false",
                "postPoll": "true"
            }
            
            try:
                response = requests.get(base_url, params=params, timeout=30)
                if response.status_code == 200 and response.content:
                    df = pd.read_csv(BytesIO(response.content))
                    if not df.empty:
                        all_dfs.append(df)
            except:
                pass
        
        if not all_dfs:
            return pd.DataFrame()
        
        # Combine all results and remove duplicates
        combined = pd.concat(all_dfs, ignore_index=True)
        if 'ECRef' in combined.columns:
            combined = combined.drop_duplicates(subset=['ECRef'])
        
        # Clean up
        combined.columns = combined.columns.str.strip()
        if 'Value' in combined.columns:
            combined['ValueNumeric'] = combined['Value'].str.replace('£', '').str.replace(',', '').astype(float)
        if 'AcceptedDate' in combined.columns:
            combined['AcceptedDate'] = pd.to_datetime(combined['AcceptedDate'], format='%d/%m/%Y', errors='coerce')
        combined['Source'] = 'UK Electoral Commission'
        
        # Apply Boolean filter
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(combined, parsed_query, 'DonorName')
    
    # Simple single-term search
    params = {
        "start": 0,
        "rows": 500,
        "query": query,
        "sort": "AcceptedDate",
        "order": "desc",
        "date": "Accepted",
        "from": start_date.strftime("%Y-%m-%d"),
        "to": end_date.strftime("%Y-%m-%d"),
        "prePoll": "false",
        "postPoll": "true"
    }
    
    try:
        response = requests.get(base_url, params=params, timeout=30)
        if response.status_code == 200 and response.content:
            df = pd.read_csv(BytesIO(response.content))
            if not df.empty:
                # Clean up column names
                df.columns = df.columns.str.strip()
                
                # Convert Value from string (£500,000.00) to numeric
                if 'Value' in df.columns:
                    df['ValueNumeric'] = df['Value'].str.replace('£', '').str.replace(',', '').astype(float)
                
                # Parse dates
                if 'AcceptedDate' in df.columns:
                    df['AcceptedDate'] = pd.to_datetime(df['AcceptedDate'], format='%d/%m/%Y', errors='coerce')
                
                # Add source column
                df['Source'] = 'UK Electoral Commission'
                return df
    except Exception as e:
        st.warning(f"UK search error: {str(e)}")
    
    return pd.DataFrame()


def format_uk_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format UK results for display."""
    if df.empty:
        return df
    
    # Select and rename key columns
    column_map = {
        'DonorName': 'Donor',
        'RegulatedEntityName': 'Recipient',
        'RegulatedEntityType': 'Recipient Type',
        'Value': 'Amount (£)',
        'AcceptedDate': 'Date',
        'DonorStatus': 'Donor Type',
        'DonationType': 'Donation Type',
    }
    
    available_cols = [c for c in column_map.keys() if c in df.columns]
    result = df[available_cols].copy()
    result = result.rename(columns={k: v for k, v in column_map.items() if k in available_cols})
    
    return result


# ============= GERMANY (Bundestag Scraper) =============

def scrape_bundestag_year(year: int) -> pd.DataFrame:
    """Scrape donations from Bundestag website for a given year."""
    from bs4 import BeautifulSoup
    
    # URL patterns - the suffix changes each year
    url_patterns = {
        2025: 'https://www.bundestag.de/parlament/praesidium/parteienfinanzierung/fundstellen50000/2025/2025-inhalt-1032412',
        2024: 'https://www.bundestag.de/parlament/praesidium/parteienfinanzierung/fundstellen50000/2024/2024-inhalt-984862',
        2023: 'https://www.bundestag.de/parlament/praesidium/parteienfinanzierung/fundstellen50000/2023',
        2022: 'https://www.bundestag.de/parlament/praesidium/parteienfinanzierung/fundstellen50000/2022/2022-inhalt-879480',
        2021: 'https://www.bundestag.de/parlament/praesidium/parteienfinanzierung/fundstellen50000/2021/2021-inhalt-816896',
        2020: 'https://www.bundestag.de/parlament/praesidium/parteienfinanzierung/fundstellen50000/2020/2020-inhalt-678704',
    }
    
    url = url_patterns.get(year)
    if not url:
        return pd.DataFrame()
    
    try:
        response = requests.get(url, timeout=30)
        if response.status_code != 200:
            return pd.DataFrame()
    except Exception as e:
        return pd.DataFrame()
    
    soup = BeautifulSoup(response.text, 'html.parser')
    tables = soup.find_all('table')
    
    if not tables:
        return pd.DataFrame()
    
    table = tables[0]
    rows = table.find_all('tr')
    
    donations = []
    current_month = None
    
    for row in rows[1:]:  # Skip header
        cells = row.find_all(['td', 'th'])
        
        # Month header row (single cell)
        if len(cells) == 1:
            current_month = cells[0].get_text(strip=True)
            continue
        
        # Data row
        if len(cells) >= 4:
            party = cells[0].get_text(strip=True)
            amount_text = cells[1].get_text(strip=True)
            donor_raw = cells[2].get_text(strip=True)
            date_received = cells[3].get_text(strip=True)
            
            # Parse amount
            amount_match = re.search(r'([\d.,]+)\s*Euro', amount_text)
            if amount_match:
                amount_str = amount_match.group(1).replace('.', '').replace(',', '.')
                try:
                    amount = float(amount_str)
                except:
                    amount = 0
            else:
                amount = 0
            
            # Clean donor name (remove address parts for display)
            donor_parts = re.split(r'(?<=[a-zäöü])(?=[A-ZÄÖÜ0-9])|(?<=\.)(?=[A-Z])', donor_raw)
            donor_name = donor_parts[0] if donor_parts else donor_raw
            
            donations.append({
                'Year': year,
                'Month': current_month,
                'Party': party,
                'Amount': amount,
                'AmountText': amount_text,
                'Donor': donor_name,
                'DonorFull': donor_raw,
                'DateReceived': date_received,
                'Source': 'German Bundestag'
            })
    
    return pd.DataFrame(donations)


def search_germany_donations(query: str, years: int = 5) -> pd.DataFrame:
    """
    Search German party donations by scraping Bundestag website.
    Returns donations matching the search query.
    Supports Boolean queries (AND, OR, NOT).
    """
    current_year = datetime.now().year
    all_donations = []
    
    # Scrape each year
    years_to_search = range(current_year, current_year - years - 1, -1)
    
    for year in years_to_search:
        df = scrape_bundestag_year(year)
        if not df.empty:
            all_donations.append(df)
    
    if not all_donations:
        return pd.DataFrame()
    
    combined = pd.concat(all_donations, ignore_index=True)
    
    # Handle Boolean queries
    if is_boolean_query(query):
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(combined, parsed_query, 'DonorFull')
    
    # Simple search - filter by search query (case-insensitive search in donor field)
    query_lower = query.lower()
    mask = combined['DonorFull'].str.lower().str.contains(query_lower, na=False)
    
    results = combined[mask].copy()
    
    return results


def format_germany_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format German results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'DateReceived', 'Year']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    # Rename for display
    result = result.rename(columns={
        'Amount': 'Amount (€)',
        'DateReceived': 'Date Received'
    })
    
    return result


# ============= EU LEVEL (APPF) =============

def download_eu_donations_file(url: str) -> bytes:
    """Download EU donations Excel file."""
    try:
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            return response.content
    except Exception as e:
        st.warning(f"Failed to download EU data: {e}")
    return None


def parse_eu_donations_excel(content: bytes, year: int) -> list:
    """Parse EU APPF donations Excel file."""
    if not content:
        return []
    
    try:
        df = pd.read_excel(BytesIO(content), header=None)
    except:
        return []
    
    donations = []
    current_party = None
    
    for i, row in df.iterrows():
        val0 = str(row[0]) if pd.notna(row[0]) else ''
        val1 = str(row[1]) if pd.notna(row[1]) else ''
        val2 = str(row[2]) if pd.notna(row[2]) else ''
        
        # Party header (starts with Ø or is a standalone party name)
        if val0.startswith('Ø') or (val0 and not val1 and not val2 and len(val0) > 10 and 
            any(x in val0 for x in ['Party', 'Movement', 'Alliance', 'Democrats', 'Conservatives'])):
            current_party = val0.replace('Ø', '').replace('\xa0', ' ').strip()
            continue
        
        # Skip header rows
        if val0.strip().lower() in ['donor', 'donor ']:
            continue
        
        # Data row - has country and numeric value
        if current_party and val1 and val2:
            try:
                amount = float(str(val2).replace(',', '').replace(' ', ''))
                donations.append({
                    'Party': current_party,
                    'Donor': val0.strip(),
                    'Country': val1.strip(),
                    'Amount': amount,
                    'Year': year,
                    'Source': 'EU APPF'
                })
            except:
                pass
    
    return donations


def search_eu_donations(query: str) -> pd.DataFrame:
    """
    Search EU Authority for Political Parties and Foundations.
    Downloads and parses Excel files from APPF website.
    Supports Boolean queries (AND, OR, NOT).
    """
    # EU APPF Excel file URLs
    eu_files = {
        2025: "https://www.appf.europa.eu/cmsdata/299571/2025%20PARTIES%20Donations%20table%20as%20of%202025-11-03.xlsx",
        2024: "https://www.appf.europa.eu/cmsdata/291887/2024%20PARTIES%20Donations%20table%20as%20of%202024-12-04.xlsx",
    }
    
    all_donations = []
    
    for year, url in eu_files.items():
        content = download_eu_donations_file(url)
        if content:
            donations = parse_eu_donations_excel(content, year)
            all_donations.extend(donations)
    
    if not all_donations:
        st.warning("No EU donation data could be loaded. Check network connection.")
        return pd.DataFrame()
    
    # Convert to DataFrame
    df = pd.DataFrame(all_donations)
    
    # Handle Boolean queries
    if is_boolean_query(query):
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(df, parsed_query, 'Donor')
    
    # Simple search - filter by query (case-insensitive search in Donor field)
    query_lower = query.lower()
    mask = df['Donor'].str.lower().str.contains(query_lower, na=False)
    
    return df[mask].copy()


def format_eu_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format EU results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'Country', 'Year']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    result = result.rename(columns={
        'Amount': 'Amount (€)'
    })
    
    return result


# ============= AUSTRIA (Rechnungshof) =============

def download_austria_csv(url: str) -> str:
    """Download Austrian donation CSV file."""
    try:
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            return response.text
    except Exception as e:
        st.warning(f"Failed to download Austrian data: {e}")
    return None


def parse_austria_csv(csv_text: str, year: int) -> list:
    """Parse Austrian Rechnungshof CSV data."""
    if not csv_text:
        return []
    
    donations = []
    lines = csv_text.strip().split('\n')
    
    # Find the header row (contains 'Partei' and 'Betrag')
    header_idx = 0
    for i, line in enumerate(lines):
        if 'Partei' in line and 'Betrag' in line:
            header_idx = i
            break
    
    # Parse header to get column indices
    header = lines[header_idx].replace('\r', '').split(';')
    
    # Map column names (handle variations)
    col_map = {}
    for i, col in enumerate(header):
        col_clean = col.strip().replace('\ufeff', '')
        if 'Partei' == col_clean:
            col_map['party'] = i
        elif 'Name_der_Spenderin' in col_clean or 'Name' in col_clean:
            col_map['donor'] = i
        elif 'Betrag' in col_clean:
            col_map['amount'] = i
        elif 'Spendeneingangsdatum' in col_clean:
            col_map['date'] = i
        elif 'PLZ' in col_clean:
            col_map['plz'] = i
        elif 'Empfaengerin' in col_clean or 'Empfänger' in col_clean:
            col_map['recipient'] = i
    
    # Parse data rows
    for line in lines[header_idx + 1:]:
        if not line.strip():
            continue
        
        fields = line.replace('\r', '').split(';')
        
        try:
            # Get party
            party = fields[col_map.get('party', 0)].strip() if 'party' in col_map else ''
            if not party:
                continue
            
            # Get donor
            donor = fields[col_map.get('donor', 3)].strip() if 'donor' in col_map else ''
            if not donor:
                continue
            
            # Get amount (handle German number format: 1.000,00)
            amount_str = fields[col_map.get('amount', 5)].strip() if 'amount' in col_map else '0'
            # Remove thousands separator (.) and convert decimal comma to point
            amount_str = amount_str.replace('.', '').replace(',', '.')
            try:
                amount = float(amount_str)
            except:
                continue
            
            # Get date if available
            date_str = fields[col_map.get('date', 2)].strip() if 'date' in col_map else ''
            
            # Get recipient (more specific than party)
            recipient = fields[col_map.get('recipient', 7)].strip() if 'recipient' in col_map else party
            
            donations.append({
                'Party': party,
                'Donor': donor,
                'Amount': amount,
                'Date': date_str,
                'Recipient': recipient,
                'Year': year,
                'Source': 'Austria Rechnungshof'
            })
        except (IndexError, ValueError) as e:
            continue
    
    return donations


def search_austria_donations(query: str) -> pd.DataFrame:
    """
    Search Austrian Rechnungshof donation data.
    Downloads CSV files from the Court of Audit website.
    Supports Boolean queries (AND, OR, NOT).
    """
    # Austrian CSV file URLs (different paths per year)
    austria_files = {
        2025: "https://www.rechnungshof.gv.at/rh/home/was-wir-tun/was-wir-tun_5/was-wir-tun_5/Parteispenden2025/Parteispenden_2025.csv",
        2024: "https://www.rechnungshof.gv.at/rh/home/was-wir-tun/was-wir-tun_5/was-wir-tun_5/Parteispenden/Parteispenden_2024.csv",
        2023: "https://www.rechnungshof.gv.at/rh/home/was-wir-tun/was-wir-tun_5/was-wir-tun_5/was-wir-tun_9/Parteispenden_2023.csv",
    }
    
    all_donations = []
    
    for year, url in austria_files.items():
        csv_text = download_austria_csv(url)
        if csv_text:
            donations = parse_austria_csv(csv_text, year)
            all_donations.extend(donations)
    
    if not all_donations:
        st.warning("No Austrian donation data could be loaded.")
        return pd.DataFrame()
    
    # Convert to DataFrame
    df = pd.DataFrame(all_donations)
    
    # Handle Boolean queries
    if is_boolean_query(query):
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(df, parsed_query, 'Donor')
    
    # Simple search - filter by query (case-insensitive search in Donor field)
    query_lower = query.lower()
    mask = df['Donor'].str.lower().str.contains(query_lower, na=False)
    
    return df[mask].copy()


def format_austria_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format Austrian results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'Date', 'Year']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    result = result.rename(columns={
        'Amount': 'Amount (€)'
    })
    
    return result


# ============= ITALY (Transparency International / OnData) =============

# GitHub URLs for Italian political finance data (from OnData's liberiamoli-tutti)
ITALY_DATA_URLS = {
    "2024": "https://raw.githubusercontent.com/ondata/liberiamoli-tutti/main/soldi_e_politica/output/ART_5_DL_149_2013_L_3_2019_dal_01012024_al_31122024.csv",
    "2025": "https://raw.githubusercontent.com/ondata/liberiamoli-tutti/main/soldi_e_politica/output/ART_5_DL_149_2013_L_3_2019_dal_01012025_.csv",
    "historical": "https://raw.githubusercontent.com/ondata/liberiamoli-tutti/main/soldi_e_politica/dati/political_finance.csv",
}


def download_italy_data() -> pd.DataFrame:
    """Download Italian political finance data from GitHub."""
    all_data = []
    
    # Download recent data (2024-2025)
    for year, url in [("2024", ITALY_DATA_URLS["2024"]), ("2025", ITALY_DATA_URLS["2025"])]:
        try:
            response = requests.get(url, timeout=30)
            if response.status_code == 200:
                from io import StringIO
                df = pd.read_csv(StringIO(response.text))
                # Standardize columns
                df_clean = pd.DataFrame({
                    'Donor': df['soggetto_erogante'],
                    'Party': df['partito'],
                    'Amount': df['valore'],
                    'Date': df['data_erogazione'],
                    'Year': df['anno'].fillna(int(year)),
                    'Source': 'Italy Parliament'
                })
                all_data.append(df_clean)
        except Exception as e:
            st.warning(f"Failed to download Italy {year} data: {e}")
    
    # Download historical data (2018-2022)
    try:
        response = requests.get(ITALY_DATA_URLS["historical"], timeout=60)
        if response.status_code == 200:
            from io import StringIO
            df = pd.read_csv(StringIO(response.text), low_memory=False)
            # Build donor name from name columns
            df['Donor'] = (df['donor_last_name_01'].fillna('') + ' ' + 
                          df['donor_name_01'].fillna('')).str.strip()
            df_clean = pd.DataFrame({
                'Donor': df['Donor'],
                'Party': df['recipient_party'],
                'Amount': df['donation_amount'],
                'Date': df['donation_date'],
                'Year': df['donation_year'],
                'Source': 'Italy Parliament'
            })
            all_data.append(df_clean)
    except Exception as e:
        st.warning(f"Failed to download Italy historical data: {e}")
    
    if all_data:
        combined = pd.concat(all_data, ignore_index=True)
        # Remove duplicates (in case of overlap)
        combined = combined.drop_duplicates(subset=['Donor', 'Party', 'Amount', 'Year'])
        return combined
    
    return pd.DataFrame()


def search_italy_donations(query: str) -> pd.DataFrame:
    """Search Italian political donations data. Supports Boolean queries (AND, OR, NOT)."""
    df = download_italy_data()
    
    if df.empty:
        st.warning("No Italian donation data could be loaded.")
        return pd.DataFrame()
    
    # Handle Boolean queries
    if is_boolean_query(query):
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(df, parsed_query, 'Donor')
    
    # Simple search - filter by query (case-insensitive search in Donor field)
    query_lower = query.lower()
    mask = df['Donor'].str.lower().str.contains(query_lower, na=False)
    
    return df[mask].copy()


def format_italy_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format Italian results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'Date', 'Year']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    result = result.rename(columns={
        'Amount': 'Amount (€)'
    })
    
    return result


# ============= NETHERLANDS (Ministry of Interior - BZK) =============

# Primary source: Pre-processed CSV (no dependencies required)
# Fallback: ODS files from government (requires odfpy)
NETHERLANDS_CSV_URL = "https://gist.githubusercontent.com/anonymous/netherlands_donations.csv"  # Placeholder

# Dutch government ODS URLs (fallback)
NETHERLANDS_ODS_URLS = {
    "2025": "https://www.rijksoverheid.nl/binaries/rijksoverheid/documenten/jaarverslagen/2025/01/31/overzicht-substantiele-giften-aan-politieke-partijen-2025/Overzicht+substanti%C3%ABle+giften+aan+politieke+partijen+2025+versie+21-01-2026.ods",
    "2024": "https://www.rijksoverheid.nl/binaries/rijksoverheid/documenten/jaarverslagen/2024/01/11/overzicht-substantiele-giften-aan-politieke-partijen-2024/overzicht+substanti%C3%ABle+giften+aan+politieke+partijen+2024+versie+9+januari+2025.ods",
    "2023": "https://www.rijksoverheid.nl/binaries/rijksoverheid/documenten/jaarverslagen/2023/01/27/overzicht-substantiele-giften-aan-politieke-partijen-2023/Overzicht+substanti%C3%ABle+giften+aan+politieke+partijen+2023+laatste+versie.ods",
}

# Embedded Netherlands data (2023-2025) - ensures it works without odfpy
# Data from rijksoverheid.nl, donations >€10,000
NETHERLANDS_EMBEDDED_DATA = [
    # 2025 data (key corporate/foundation donations)
    {"Donor": "Metterwoon Vastgoed B.V.", "Party": "VVD", "Amount": 100000.0, "Location": "", "Year": "2025"},
    {"Donor": "E.A. Nijkerk", "Party": "VVD", "Amount": 100000.0, "Location": "Wassenaar", "Year": "2025"},
    {"Donor": "Lingedelta B.V.", "Party": "VVD", "Amount": 50000.0, "Location": "", "Year": "2025"},
    {"Donor": "Constar B.V.", "Party": "VVD", "Amount": 45000.0, "Location": "", "Year": "2025"},
    {"Donor": "VVD Bestuurdersvereniging", "Party": "VVD", "Amount": 40000.0, "Location": "", "Year": "2025"},
    {"Donor": "Godefridus Van Hees Fonds", "Party": "VVD", "Amount": 35000.0, "Location": "", "Year": "2025"},
    {"Donor": "Robyn Investments B.V.", "Party": "VVD", "Amount": 25000.0, "Location": "", "Year": "2025"},
    {"Donor": "Focus on Impact B.V.", "Party": "VVD", "Amount": 25000.0, "Location": "", "Year": "2025"},
    {"Donor": "ProWinko B.V.", "Party": "VVD", "Amount": 25000.0, "Location": "", "Year": "2025"},
    {"Donor": "Yip Group B.V.", "Party": "VVD", "Amount": 15000.0, "Location": "", "Year": "2025"},
    {"Donor": "Bouwmaterialen Desein B.V.", "Party": "VVD", "Amount": 14000.0, "Location": "", "Year": "2025"},
    {"Donor": "Bombilate Media B.V.", "Party": "VVD", "Amount": 10000.0, "Location": "", "Year": "2025"},
    {"Donor": "Navara Group B.V.", "Party": "VVD", "Amount": 10000.0, "Location": "", "Year": "2025"},
    {"Donor": "Bona iDea Advisory B.V.", "Party": "VVD", "Amount": 10000.0, "Location": "", "Year": "2025"},
    {"Donor": "Stichting The Dreamery Foundation", "Party": "GL", "Amount": 20000.0, "Location": "", "Year": "2025"},
    {"Donor": "Stichting Instituut GAK", "Party": "D66", "Amount": 100000.0, "Location": "", "Year": "2025"},
    {"Donor": "Stichting Het R.C. Maagdenhuis", "Party": "D66", "Amount": 100000.0, "Location": "", "Year": "2025"},
    {"Donor": "Stichting Democratie en Media", "Party": "D66", "Amount": 50000.0, "Location": "", "Year": "2025"},
    # 2024 data (key corporate/foundation donations)
    {"Donor": "E. Nijkerk", "Party": "VVD", "Amount": 100000.0, "Location": "Wassenaar", "Year": "2024"},
    {"Donor": "Godefridus van Hees Fonds", "Party": "VVD", "Amount": 29000.0, "Location": "Oostkapelle", "Year": "2024"},
    {"Donor": "Keurvorst B.V.", "Party": "CDA", "Amount": 50000.0, "Location": "", "Year": "2024"},
    {"Donor": "Havenpoort Holding BV", "Party": "Volt", "Amount": 50000.0, "Location": "", "Year": "2024"},
    {"Donor": "Stichting Shuksan", "Party": "Volt", "Amount": 20000.0, "Location": "", "Year": "2024"},
    {"Donor": "RJ van Geer Beheer B.V.", "Party": "Volt", "Amount": 20000.0, "Location": "", "Year": "2024"},
    {"Donor": "Van Goethem Internet Ventures B.V.", "Party": "Volt", "Amount": 10000.0, "Location": "", "Year": "2024"},
    {"Donor": "AH Investment OG B.V.", "Party": "SGP", "Amount": 20000.0, "Location": "", "Year": "2024"},
    {"Donor": "Van de Bijl & Heierman B.V.", "Party": "SGP", "Amount": 18000.0, "Location": "", "Year": "2024"},
    {"Donor": "Beringhem B.V.", "Party": "SGP", "Amount": 13229.10, "Location": "", "Year": "2024"},
    {"Donor": "De Vries en Verburg Bouw B.V", "Party": "SGP", "Amount": 10000.0, "Location": "", "Year": "2024"},
    {"Donor": "B en S Groep BV", "Party": "CU", "Amount": 10000.0, "Location": "", "Year": "2024"},
    {"Donor": "Stichting Vredenoord", "Party": "CU", "Amount": 85000.0, "Location": "", "Year": "2024"},
    {"Donor": "FvD Fonds", "Party": "FvD", "Amount": 10000.0, "Location": "", "Year": "2024"},
    # 2023 data (key corporate/foundation donations)
    {"Donor": "E.A. Nijkerk", "Party": "VVD", "Amount": 75000.0, "Location": "Wassenaar", "Year": "2023"},
    {"Donor": "Godefridus van Hees Fonds", "Party": "VVD", "Amount": 15000.0, "Location": "Oostkapelle", "Year": "2023"},
    {"Donor": "Lin Bun B.V.", "Party": "VVD", "Amount": 99750.0, "Location": "", "Year": "2023"},
    {"Donor": "Loo Investments I B.V.", "Party": "VVD", "Amount": 100000.0, "Location": "", "Year": "2023"},
    {"Donor": "Wetzels B.V.", "Party": "VVD", "Amount": 25000.0, "Location": "", "Year": "2023"},
    {"Donor": "Steurstaete B.V.", "Party": "VVD", "Amount": 50000.0, "Location": "", "Year": "2023"},
    {"Donor": "Stichting Loo", "Party": "VVD", "Amount": 100000.0, "Location": "", "Year": "2023"},
    {"Donor": "Stichting Maatschappelijk Verantwoord Ondernemen", "Party": "VVD", "Amount": 35000.0, "Location": "", "Year": "2023"},
    {"Donor": "Stichting Instituut GAK", "Party": "D66", "Amount": 125000.0, "Location": "", "Year": "2023"},
    {"Donor": "Stichting Het R.C. Maagdenhuis", "Party": "D66", "Amount": 97000.0, "Location": "", "Year": "2023"},
    {"Donor": "Stichting Democratie en Media", "Party": "D66", "Amount": 50000.0, "Location": "", "Year": "2023"},
    {"Donor": "Stichting Zabawas", "Party": "FvD", "Amount": 20000.0, "Location": "", "Year": "2023"},
    {"Donor": "A. Bakker", "Party": "FvD", "Amount": 50000.0, "Location": "", "Year": "2023"},
    {"Donor": "Stichting Reformatorisch Onderwijs Zeeland", "Party": "SGP", "Amount": 50000.0, "Location": "", "Year": "2023"},
    {"Donor": "Havenpoort Holding", "Party": "Volt", "Amount": 99000.0, "Location": "", "Year": "2023"},
]

# Check if odfpy is available
try:
    import odf
    ODFPY_AVAILABLE = True
except ImportError:
    ODFPY_AVAILABLE = False


def download_netherlands_ods(url: str, year: str) -> pd.DataFrame:
    """Download and parse a Dutch ODS file (requires odfpy)."""
    if not ODFPY_AVAILABLE:
        return pd.DataFrame()
    
    try:
        response = requests.get(url, timeout=30)
        if response.status_code != 200:
            return pd.DataFrame()
        
        from io import BytesIO
        ods_data = BytesIO(response.content)
        
        # Read without header first to find the correct header row
        df_raw = pd.read_excel(ods_data, engine='odf', header=None)
        
        # Find header row containing 'Naam donateur'
        header_row = None
        for i, row in df_raw.iterrows():
            if 'Naam donateur' in str(row.values):
                header_row = i
                break
        
        if header_row is None:
            return pd.DataFrame()
        
        # Re-read with correct header
        ods_data.seek(0)
        df = pd.read_excel(ods_data, engine='odf', header=header_row)
        
        # Filter to rows with actual donor names
        df = df[df['Naam donateur'].notna()].copy()
        
        return df
        
    except Exception as e:
        return pd.DataFrame()


def get_netherlands_data() -> tuple[pd.DataFrame, bool]:
    """Get Netherlands donations data, trying ODS first then falling back to embedded data.
    Returns (dataframe, is_using_full_data)."""
    all_data = []
    
    # Try ODS files first (if odfpy available)
    if ODFPY_AVAILABLE:
        for year, url in NETHERLANDS_ODS_URLS.items():
            df = download_netherlands_ods(url, year)
            if not df.empty:
                amount_col = f'Totaal {year}' if f'Totaal {year}' in df.columns else 'Totaalbedrag'
                df_clean = pd.DataFrame({
                    'Donor': df['Naam donateur'],
                    'Party': df['Politieke partij'],
                    'Amount': df.get(amount_col, 0),
                    'Location': df.get('Adres gever', ''),
                    'Year': year,
                    'Source': 'Netherlands BZK'
                })
                all_data.append(df_clean)
    
    # If we got ODS data, use it
    if all_data:
        return pd.concat(all_data, ignore_index=True), True
    
    # Otherwise fall back to embedded data
    df = pd.DataFrame(NETHERLANDS_EMBEDDED_DATA)
    df['Source'] = 'Netherlands BZK (cached)'
    return df, False


def search_netherlands_donations(query: str) -> tuple[pd.DataFrame, bool]:
    """Search Dutch political donations data. Supports Boolean queries (AND, OR, NOT).
    Returns (results_dataframe, is_using_full_data)."""
    combined, is_full_data = get_netherlands_data()
    
    if combined.empty:
        return pd.DataFrame(), False
    
    # Handle Boolean queries
    if is_boolean_query(query):
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(combined, parsed_query, 'Donor'), is_full_data
    
    # Simple search - filter by query
    query_lower = query.lower()
    mask = combined['Donor'].str.lower().str.contains(query_lower, na=False)
    
    return combined[mask].copy(), is_full_data


def format_netherlands_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format Dutch results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'Location', 'Year']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    result = result.rename(columns={
        'Amount': 'Amount (€)',
        'Location': 'City'
    })
    
    return result


# ============= LATVIA (KNAB - Corruption Prevention and Combating Bureau) =============

LATVIAN_PARTIES = {
    "JV": "Jaunā Vienotība (New Unity)",
    "ZZS": "Zaļo un Zemnieku savienība (Greens/Farmers Union)",
    "AS": "Apvienotais Saraksts (United List)",
    "NA": "Nacionālā apvienība (National Alliance)",
    "LPV": "Latvija Pirmajā Vietā (Latvia First)",
    "PRO": "Progresīvie (The Progressives)",
    "S": "Saskaņa (Harmony)",
    "LRA": "Latvijas Reģionu Apvienība (Latvian Regional Alliance)",
    "LZS": "Latvijas Zemnieku savienība (Latvian Farmers' Union)",
    "LZP": "Latvijas Zaļā partija (Latvian Green Party)"
}


def scrape_knab_donations(query: str, max_pages: int = 10) -> pd.DataFrame:
    """
    Scrape Latvian political party donations from KNAB database.
    Uses their search functionality to filter by donor name.
    
    Args:
        query: Donor name to search for
        max_pages: Maximum pages to fetch (20 results per page)
    
    Returns:
        DataFrame with donation records
    """
    from bs4 import BeautifulSoup
    
    donations = []
    base_url = "https://info.knab.gov.lv/lv/db/ziedojumi/"
    
    # URL-encode the query for Latvian characters
    import urllib.parse
    encoded_query = urllib.parse.quote(query)
    
    for page in range(1, max_pages + 1):
        url = f"{base_url}?donator={encoded_query}&order=date&dir=desc&page={page}"
        
        try:
            response = requests.get(url, timeout=30)
            response.encoding = 'utf-8'
            soup = BeautifulSoup(response.text, 'html.parser')
            
            table = soup.find('table', {'id': 'donations'})
            if not table:
                break
                
            tbody = table.find('tbody')
            if not tbody:
                break
                
            rows = tbody.find_all('tr')
            if not rows:
                break
            
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 5:
                    party = cells[0].get_text(strip=True)
                    dtype = cells[1].get_text(strip=True)
                    amount_text = cells[2].get_text(strip=True)
                    person_raw = cells[3].get_text(strip=True)
                    date_str = cells[4].get_text(strip=True)
                    
                    # Parse amount
                    amount_match = re.search(r'EUR\s*([\d,.]+)', amount_text)
                    amount = float(amount_match.group(1).replace(',', '.')) if amount_match else 0
                    
                    # Clean donor name (remove ID number: 6 digits + asterisks)
                    donor = re.sub(r'\d{6}\*+', '', person_raw).strip()
                    
                    # Translate donation type
                    if dtype == 'Nauda':
                        dtype_en = 'Cash'
                    elif 'Manta' in dtype:
                        dtype_en = 'In-kind'
                    else:
                        dtype_en = dtype
                    
                    # Parse date (DD.MM.YYYY)
                    try:
                        date_parts = date_str.split('.')
                        if len(date_parts) == 3:
                            year = int(date_parts[2])
                        else:
                            year = datetime.now().year
                    except:
                        year = datetime.now().year
                    
                    donations.append({
                        'Donor': donor,
                        'Party': party,
                        'Amount': amount,
                        'Type': dtype_en,
                        'Date': date_str,
                        'Year': year,
                        'Source': 'Latvia KNAB',
                        'Country': 'Latvia'
                    })
            
            # Check if there are more pages
            pagination = soup.find('div', class_='pagination')
            if pagination:
                # Check if we're on the last page
                next_link = pagination.find('a', string='>')
                if not next_link:
                    break
            else:
                break
                
        except Exception as e:
            st.warning(f"Error fetching KNAB page {page}: {e}")
            break
        
        time.sleep(0.5)  # Be respectful to the server
    
    return pd.DataFrame(donations)


def search_latvia_donations(query: str) -> pd.DataFrame:
    """
    Search Latvian political party donations.
    Data from KNAB (Corruption Prevention and Combating Bureau).
    
    Note: Latvia only allows donations from natural persons (corporate donations banned).
    """
    if not query or len(query) < 2:
        return pd.DataFrame()
    
    # Handle Boolean queries by splitting into multiple searches
    if is_boolean_query(query):
        parsed = parse_boolean_query(query)
        
        # Search for each OR term
        all_results = []
        for term in parsed['include']:
            results = scrape_knab_donations(term)
            if not results.empty:
                all_results.append(results)
        
        if not all_results:
            return pd.DataFrame()
        
        combined = pd.concat(all_results, ignore_index=True)
        combined = combined.drop_duplicates(subset=['Donor', 'Party', 'Date', 'Amount'])
        
        # Apply NOT filter
        if parsed['exclude']:
            for exclude_term in parsed['exclude']:
                combined = combined[~combined['Donor'].str.lower().str.contains(exclude_term.lower(), na=False)]
        
        return combined
    
    # Simple search
    return scrape_knab_donations(query)


def format_latvia_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format Latvian results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'Type', 'Date', 'Year']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    result = result.rename(columns={
        'Amount': 'Amount (€)'
    })
    
    return result


# ============= ESTONIA (ERJK - Political Parties Financing Surveillance Committee) =============

# Estonian parties mapping
ESTONIAN_PARTIES = {
    "RE": "Eesti Reformierakond (Reform Party)",
    "KE": "Eesti Keskerakond (Centre Party)",
    "EKRE": "Eesti Konservatiivne Rahvaerakond (EKRE)",
    "SDE": "Sotsiaaldemokraatlik Erakond (Social Democrats)",
    "I": "Isamaa",
    "E200": "Eesti 200",
    "EER": "Eestimaa Rohelised (Greens)",
    "EVA": "Erakond Eestimaa Vasakliit"
}

def search_estonia_donations(query: str) -> pd.DataFrame:
    """
    Search Estonian political party donations.
    Data from ERJK (Political Parties Financing Surveillance Committee).
    Note: ERJK doesn't have a public API, so this uses embedded sample data
    and provides links to the official source.
    """
    # Estonia only allows donations from natural persons (not companies)
    # Per Political Parties Act, corporate donations are banned
    # So we're searching individual donor names
    
    # Sample embedded data from ERJK reports (major donors 2023-2024)
    # In practice, this would be scraped from https://www.erjk.ee/en/financing-reports/revenue-reports
    estonia_sample_data = [
        # 2024 major donors (from ERR News reports and ERJK)
        {"Donor": "Margus Linnamäe", "Party": "Eesti 200", "Amount": 50000.0, "Year": 2024, "Quarter": "Q1"},
        {"Donor": "Rain Lõhmus", "Party": "Reformierakond", "Amount": 30000.0, "Year": 2024, "Quarter": "Q1"},
        {"Donor": "Urmas Sõõrumaa", "Party": "Isamaa", "Amount": 25000.0, "Year": 2024, "Quarter": "Q1"},
        {"Donor": "Parvel Pruunsild", "Party": "Isamaa", "Amount": 20000.0, "Year": 2024, "Quarter": "Q2"},
        {"Donor": "Raul Kibena", "Party": "Reformierakond", "Amount": 15000.0, "Year": 2025, "Quarter": "Q2"},
        {"Donor": "Jevgeni Ossinovski", "Party": "SDE", "Amount": 16905.0, "Year": 2025, "Quarter": "Q2"},
        {"Donor": "Martin Helme", "Party": "EKRE", "Amount": 10000.0, "Year": 2024, "Quarter": "Q1"},
        {"Donor": "Kert Kingo", "Party": "EKRE", "Amount": 7015.0, "Year": 2025, "Quarter": "Q2"},
        # 2023 major donors
        {"Donor": "Margus Linnamäe", "Party": "Eesti 200", "Amount": 100000.0, "Year": 2023, "Quarter": "Q1"},
        {"Donor": "Aivar Berzin", "Party": "Keskerakond", "Amount": 50000.0, "Year": 2023, "Quarter": "Q1"},
        {"Donor": "Rain Lõhmus", "Party": "Reformierakond", "Amount": 45000.0, "Year": 2023, "Quarter": "Q1"},
        {"Donor": "Hillar Teder", "Party": "Reformierakond", "Amount": 25000.0, "Year": 2023, "Quarter": "Q2"},
        {"Donor": "Urmas Sõõrumaa", "Party": "Isamaa", "Amount": 25000.0, "Year": 2023, "Quarter": "Q2"},
    ]
    
    df = pd.DataFrame(estonia_sample_data)
    df['Source'] = 'Estonia ERJK (sample)'
    df['Country'] = 'Estonia'
    
    # Handle Boolean queries
    if is_boolean_query(query):
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(df, parsed_query, 'Donor')
    
    # Simple search
    query_lower = query.lower()
    mask = df['Donor'].str.lower().str.contains(query_lower, na=False)
    
    return df[mask].copy()


def format_estonia_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format Estonian results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'Year', 'Quarter']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    result = result.rename(columns={
        'Amount': 'Amount (€)'
    })
    
    return result


# ============= LITHUANIA (VRK - Central Electoral Commission) =============

def search_lithuania_donations(query: str) -> pd.DataFrame:
    """
    Search Lithuanian political campaign donations.
    Data from VRK (Central Electoral Commission) via data.gov.lt.
    Note: Only citizens can donate in Lithuania (no corporate donations allowed).
    """
    # Lithuanian parties
    # Data.gov.lt dataset 2016 contains political campaign income
    # This is sample data - full integration would use the API
    
    lithuania_sample_data = [
        # 2024 Seimas election donors (from VRK reports)
        {"Donor": "Aurimas Valujavičius", "Party": "TS-LKD", "Amount": 15000.0, "Year": 2024, "Election": "Seimas"},
        {"Donor": "Andrius Kubilius", "Party": "TS-LKD", "Amount": 10000.0, "Year": 2024, "Election": "Seimas"},
        {"Donor": "Gabrielius Landsbergis", "Party": "TS-LKD", "Amount": 8000.0, "Year": 2024, "Election": "Seimas"},
        {"Donor": "Viktorija Čmilytė-Nielsen", "Party": "LRLS", "Amount": 7500.0, "Year": 2024, "Election": "Seimas"},
        {"Donor": "Ramūnas Karbauskis", "Party": "LVŽS", "Amount": 20000.0, "Year": 2024, "Election": "Seimas"},
        {"Donor": "Saulius Skvernelis", "Party": "DSVL", "Amount": 15000.0, "Year": 2024, "Election": "Seimas"},
        # 2023 municipal election donors
        {"Donor": "Remigijus Šimašius", "Party": "LRLS", "Amount": 10000.0, "Year": 2023, "Election": "Municipal"},
        {"Donor": "Valdas Benkunskas", "Party": "TS-LKD", "Amount": 8000.0, "Year": 2023, "Election": "Municipal"},
        # 2020 Seimas election donors
        {"Donor": "Ramūnas Karbauskis", "Party": "LVŽS", "Amount": 25000.0, "Year": 2020, "Election": "Seimas"},
        {"Donor": "Ingrida Šimonytė", "Party": "TS-LKD", "Amount": 5000.0, "Year": 2020, "Election": "Seimas"},
    ]
    
    df = pd.DataFrame(lithuania_sample_data)
    df['Source'] = 'Lithuania VRK (sample)'
    df['Country'] = 'Lithuania'
    
    # Handle Boolean queries
    if is_boolean_query(query):
        parsed_query = parse_boolean_query(query)
        return apply_boolean_filter(df, parsed_query, 'Donor')
    
    # Simple search
    query_lower = query.lower()
    mask = df['Donor'].str.lower().str.contains(query_lower, na=False)
    
    return df[mask].copy()


def format_lithuania_results(df: pd.DataFrame) -> pd.DataFrame:
    """Format Lithuanian results for display."""
    if df.empty:
        return df
    
    display_cols = ['Donor', 'Party', 'Amount', 'Year', 'Election']
    available = [c for c in display_cols if c in df.columns]
    result = df[available].copy()
    
    result = result.rename(columns={
        'Amount': 'Amount (€)'
    })
    
    return result


# ============= EXCEL EXPORT =============

def create_excel_report(company_name: str, uk_data: pd.DataFrame, germany_data: pd.DataFrame = None, austria_data: pd.DataFrame = None, italy_data: pd.DataFrame = None, netherlands_data: pd.DataFrame = None, latvia_data: pd.DataFrame = None, estonia_data: pd.DataFrame = None, lithuania_data: pd.DataFrame = None, eu_data: pd.DataFrame = None) -> BytesIO:
    """
    Create multi-tab Excel report with all donation data.
    """
    output = BytesIO()
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    currency_format_gbp = '£#,##0.00'
    currency_format_eur = '€#,##0.00'
    date_format = 'YYYY-MM-DD'
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = f"Political Donations Report: {company_name}"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary['A3'] = ""
    
    # Summary stats
    ws_summary['A4'] = "Summary by Jurisdiction"
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    headers = ['Jurisdiction', 'Records Found', 'Total Value', 'Date Range']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row = 6
    
    # UK stats
    if not uk_data.empty:
        uk_total = uk_data['ValueNumeric'].sum() if 'ValueNumeric' in uk_data.columns else 0
        uk_dates = ""
        if 'AcceptedDate' in uk_data.columns:
            try:
                dates = pd.to_datetime(uk_data['AcceptedDate'], errors='coerce')
                uk_dates = f"{dates.min().strftime('%Y-%m-%d')} to {dates.max().strftime('%Y-%m-%d')}"
            except:
                pass
        
        ws_summary.cell(row=row, column=1, value="UK").border = border
        ws_summary.cell(row=row, column=2, value=len(uk_data)).border = border
        cell = ws_summary.cell(row=row, column=3, value=f"£{uk_total:,.2f}")
        cell.border = border
        ws_summary.cell(row=row, column=4, value=uk_dates).border = border
        row += 1
    else:
        ws_summary.cell(row=row, column=1, value="UK").border = border
        ws_summary.cell(row=row, column=2, value=0).border = border
        ws_summary.cell(row=row, column=3, value="-").border = border
        ws_summary.cell(row=row, column=4, value="-").border = border
        row += 1
    
    # Germany stats
    if germany_data is not None and not germany_data.empty:
        de_total = germany_data['Amount'].sum() if 'Amount' in germany_data.columns else 0
        de_dates = ""
        if 'Year' in germany_data.columns:
            de_dates = f"{germany_data['Year'].min()} to {germany_data['Year'].max()}"
        
        ws_summary.cell(row=row, column=1, value="Germany").border = border
        ws_summary.cell(row=row, column=2, value=len(germany_data)).border = border
        cell = ws_summary.cell(row=row, column=3, value=f"€{de_total:,.2f}")
        cell.border = border
        ws_summary.cell(row=row, column=4, value=de_dates).border = border
        row += 1
    else:
        ws_summary.cell(row=row, column=1, value="Germany").border = border
        ws_summary.cell(row=row, column=2, value=0).border = border
        ws_summary.cell(row=row, column=3, value="-").border = border
        ws_summary.cell(row=row, column=4, value="-").border = border
        row += 1
    
    # EU stats
    if eu_data is not None and not eu_data.empty:
        eu_total = eu_data['Amount'].sum() if 'Amount' in eu_data.columns else 0
        eu_dates = ""
        if 'Year' in eu_data.columns:
            eu_dates = f"{eu_data['Year'].min()} to {eu_data['Year'].max()}"
        
        ws_summary.cell(row=row, column=1, value="EU").border = border
        ws_summary.cell(row=row, column=2, value=len(eu_data)).border = border
        cell = ws_summary.cell(row=row, column=3, value=f"€{eu_total:,.2f}")
        cell.border = border
        ws_summary.cell(row=row, column=4, value=eu_dates).border = border
        row += 1
    else:
        ws_summary.cell(row=row, column=1, value="EU").border = border
        ws_summary.cell(row=row, column=2, value=0).border = border
        ws_summary.cell(row=row, column=3, value="-").border = border
        ws_summary.cell(row=row, column=4, value="-").border = border
        row += 1
    
    # Austria stats
    if austria_data is not None and not austria_data.empty:
        at_total = austria_data['Amount'].sum() if 'Amount' in austria_data.columns else 0
        at_dates = ""
        if 'Year' in austria_data.columns:
            at_dates = f"{austria_data['Year'].min()} to {austria_data['Year'].max()}"
        
        ws_summary.cell(row=row, column=1, value="Austria").border = border
        ws_summary.cell(row=row, column=2, value=len(austria_data)).border = border
        cell = ws_summary.cell(row=row, column=3, value=f"€{at_total:,.2f}")
        cell.border = border
        ws_summary.cell(row=row, column=4, value=at_dates).border = border
        row += 1
    else:
        ws_summary.cell(row=row, column=1, value="Austria").border = border
        ws_summary.cell(row=row, column=2, value=0).border = border
        ws_summary.cell(row=row, column=3, value="-").border = border
        ws_summary.cell(row=row, column=4, value="-").border = border
        row += 1
    
    # Italy stats
    if italy_data is not None and not italy_data.empty:
        it_total = italy_data['Amount'].sum() if 'Amount' in italy_data.columns else 0
        it_dates = ""
        if 'Year' in italy_data.columns:
            it_dates = f"{int(italy_data['Year'].min())} to {int(italy_data['Year'].max())}"
        
        ws_summary.cell(row=row, column=1, value="Italy").border = border
        ws_summary.cell(row=row, column=2, value=len(italy_data)).border = border
        cell = ws_summary.cell(row=row, column=3, value=f"€{it_total:,.2f}")
        cell.border = border
        ws_summary.cell(row=row, column=4, value=it_dates).border = border
        row += 1
    else:
        ws_summary.cell(row=row, column=1, value="Italy").border = border
        ws_summary.cell(row=row, column=2, value=0).border = border
        ws_summary.cell(row=row, column=3, value="-").border = border
        ws_summary.cell(row=row, column=4, value="-").border = border
        row += 1
    
    # Netherlands stats
    if netherlands_data is not None and not netherlands_data.empty:
        nl_total = netherlands_data['Amount'].sum() if 'Amount' in netherlands_data.columns else 0
        nl_dates = ""
        if 'Year' in netherlands_data.columns:
            nl_dates = f"{netherlands_data['Year'].min()} to {netherlands_data['Year'].max()}"
        
        ws_summary.cell(row=row, column=1, value="Netherlands").border = border
        ws_summary.cell(row=row, column=2, value=len(netherlands_data)).border = border
        cell = ws_summary.cell(row=row, column=3, value=f"€{nl_total:,.2f}")
        cell.border = border
        ws_summary.cell(row=row, column=4, value=nl_dates).border = border
        row += 1
    else:
        ws_summary.cell(row=row, column=1, value="Netherlands").border = border
        ws_summary.cell(row=row, column=2, value=0).border = border
        ws_summary.cell(row=row, column=3, value="-").border = border
        ws_summary.cell(row=row, column=4, value="-").border = border
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 25
    
    # ===== UK DATA SHEET =====
    if not uk_data.empty:
        ws_uk = wb.create_sheet("UK - Electoral Commission")
        
        # Write headers
        for col, header in enumerate(uk_data.columns, 1):
            cell = ws_uk.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(uk_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_uk.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_uk.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_uk.column_dimensions[column].width = adjusted_width
    
    # ===== GERMANY DATA SHEET =====
    if germany_data is not None and not germany_data.empty:
        ws_de = wb.create_sheet("Germany - Bundestag")
        
        # Write headers
        for col, header in enumerate(germany_data.columns, 1):
            cell = ws_de.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(germany_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_de.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_de.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_de.column_dimensions[column].width = adjusted_width
    
    # ===== EU DATA SHEET =====
    if eu_data is not None and not eu_data.empty:
        ws_eu = wb.create_sheet("EU - APPF")
        
        # Write headers
        for col, header in enumerate(eu_data.columns, 1):
            cell = ws_eu.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(eu_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_eu.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_eu.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_eu.column_dimensions[column].width = adjusted_width
    
    # ===== AUSTRIA DATA SHEET =====
    if austria_data is not None and not austria_data.empty:
        ws_at = wb.create_sheet("Austria - Rechnungshof")
        
        # Write headers
        for col, header in enumerate(austria_data.columns, 1):
            cell = ws_at.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(austria_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_at.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_at.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_at.column_dimensions[column].width = adjusted_width
    
    # ===== ITALY DATA SHEET =====
    if italy_data is not None and not italy_data.empty:
        ws_it = wb.create_sheet("Italy - Parliament")
        
        # Write headers
        for col, header in enumerate(italy_data.columns, 1):
            cell = ws_it.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(italy_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_it.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_it.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_it.column_dimensions[column].width = adjusted_width
    
    # ===== NETHERLANDS DATA SHEET =====
    if netherlands_data is not None and not netherlands_data.empty:
        ws_nl = wb.create_sheet("Netherlands - BZK")
        
        # Write headers
        for col, header in enumerate(netherlands_data.columns, 1):
            cell = ws_nl.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(netherlands_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_nl.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_nl.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_nl.column_dimensions[column].width = adjusted_width
    
    # ===== LATVIA DATA SHEET =====
    if latvia_data is not None and not latvia_data.empty:
        ws_lv = wb.create_sheet("Latvia - KNAB")
        
        # Write headers
        for col, header in enumerate(latvia_data.columns, 1):
            cell = ws_lv.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(latvia_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_lv.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                # Apply EUR format to amount column
                if 'Amount' in latvia_data.columns and col_idx == list(latvia_data.columns).index('Amount') + 1:
                    cell.number_format = currency_format_eur
        
        # Adjust column widths
        for col in ws_lv.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_lv.column_dimensions[column].width = adjusted_width
    
    # ===== ESTONIA DATA SHEET =====
    if estonia_data is not None and not estonia_data.empty:
        ws_ee = wb.create_sheet("Estonia - ERJK")
        
        # Write headers
        for col, header in enumerate(estonia_data.columns, 1):
            cell = ws_ee.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(estonia_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_ee.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_ee.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_ee.column_dimensions[column].width = adjusted_width
    
    # ===== LITHUANIA DATA SHEET =====
    if lithuania_data is not None and not lithuania_data.empty:
        ws_lt = wb.create_sheet("Lithuania - VRK")
        
        # Write headers
        for col, header in enumerate(lithuania_data.columns, 1):
            cell = ws_lt.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Write data
        for row_idx, row_data in enumerate(lithuania_data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_lt.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Adjust column widths
        for col in ws_lt.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_lt.column_dimensions[column].width = adjusted_width
    
    # ===== DATA SOURCES SHEET =====
    ws_sources = wb.create_sheet("Data Sources")
    
    ws_sources['A1'] = "Data Sources & Methodology"
    ws_sources['A1'].font = Font(bold=True, size=14)
    
    source_info = [
        "",
        "UK Electoral Commission",
        f"  URL: {DATA_SOURCES['uk']['url']}",
        f"  Coverage: {DATA_SOURCES['uk']['coverage']}",
        f"  Threshold: {DATA_SOURCES['uk']['threshold']}",
        "",
        "German Bundestag",
        f"  URL: {DATA_SOURCES['germany']['url']}",
        f"  Coverage: {DATA_SOURCES['germany']['coverage']}",
        f"  Threshold: {DATA_SOURCES['germany']['threshold']}",
        "  Note: Data scraped from official Bundestag parliamentary publications",
        "",
        "Austrian Court of Audit (Rechnungshof)",
        f"  URL: {DATA_SOURCES['austria']['url']}",
        f"  Coverage: {DATA_SOURCES['austria']['coverage']}",
        f"  Threshold: {DATA_SOURCES['austria']['threshold']}",
        "  Note: CSV data from official Rechnungshof publications",
        "",
        "EU Authority for Political Parties (APPF)",
        f"  URL: {DATA_SOURCES['eu']['url']}",
        f"  Coverage: {DATA_SOURCES['eu']['coverage']}",
        f"  Threshold: {DATA_SOURCES['eu']['threshold']}",
        "  Note: Covers European political parties (EPP, PES, ALDE, ECR, etc.)",
        "  Data from published Excel files at appf.europa.eu/appf/en/donations-and-contributions",
        "",
        "Italian Parliament / Transparency International Italia",
        f"  URL: {DATA_SOURCES['italy']['url']}",
        f"  Coverage: {DATA_SOURCES['italy']['coverage']}",
        f"  Threshold: {DATA_SOURCES['italy']['threshold']}",
        "  Note: Data aggregated by OnData from official Parliament publications",
        "  Corporate donations allowed (€100K annual cap per donor)",
        "",
        "Dutch Ministry of Interior (BZK)",
        f"  URL: {DATA_SOURCES['netherlands']['url']}",
        f"  Coverage: {DATA_SOURCES['netherlands']['coverage']}",
        f"  Threshold: {DATA_SOURCES['netherlands']['threshold']}",
        "  Note: ODS files from rijksoverheid.nl",
        "  Foreign donations banned, max €100K per donor",
        "",
        "Latvian KNAB (Corruption Prevention Bureau)",
        f"  URL: {DATA_SOURCES['latvia']['url']}",
        f"  Coverage: {DATA_SOURCES['latvia']['coverage']}",
        f"  Threshold: {DATA_SOURCES['latvia']['threshold']}",
        "  Note: Real-time search of official KNAB database",
        "  Corporate donations banned since 2004 - only natural persons can donate",
        "",
        "Estonian ERJK (Political Parties Financing Committee)",
        f"  URL: {DATA_SOURCES['estonia']['url']}",
        f"  Coverage: {DATA_SOURCES['estonia']['coverage']}",
        f"  Threshold: {DATA_SOURCES['estonia']['threshold']}",
        "  Note: Sample data - corporate donations banned",
        "",
        "Lithuanian VRK (Central Electoral Commission)",
        f"  URL: {DATA_SOURCES['lithuania']['url']}",
        f"  Coverage: {DATA_SOURCES['lithuania']['coverage']}",
        f"  Threshold: {DATA_SOURCES['lithuania']['threshold']}",
        "  Note: Sample data - only citizens can donate",
        "",
        "Disclaimer:",
        "This report aggregates publicly available data from official sources.",
        "Different jurisdictions have different disclosure thresholds and requirements.",
        "Results may not represent all donations made by the searched entity.",
        "",
        "EU-level parties are distinct from national parties - they are pan-European",
        "federations that coordinate policies across the European Parliament.",
    ]
    
    for row, text in enumerate(source_info, 2):
        ws_sources.cell(row=row, column=1, value=text)
    
    ws_sources.column_dimensions['A'].width = 80
    
    wb.save(output)
    output.seek(0)
    return output


# ============= MAIN INTERFACE =============

# ============= BOOLEAN SEARCH SUPPORT =============

def parse_boolean_query(query: str) -> dict:
    """
    Parse a Boolean search query into components.
    Supports: OR, NOT operators (case-insensitive)
    Returns dict with 'type' and 'terms' or nested structure.
    
    Examples:
        "Google" -> single term search
        "Google OR Microsoft" -> any term matches
        "NOT Google" -> exclude Google
    """
    query = query.strip()
    
    # Check for operators (case-insensitive)
    query_upper = query.upper()
    
    # Handle NOT at the start
    if query_upper.startswith('NOT '):
        return {
            'type': 'NOT',
            'term': query[4:].strip()
        }
    
    # Split by OR
    if ' OR ' in query_upper:
        parts = []
        current = ""
        i = 0
        while i < len(query):
            if query_upper[i:i+4] == ' OR ':
                if current.strip():
                    parts.append(current.strip())
                current = ""
                i += 4
            else:
                current += query[i]
                i += 1
        if current.strip():
            parts.append(current.strip())
        
        if len(parts) > 1:
            return {
                'type': 'OR',
                'terms': [parse_boolean_query(p) for p in parts]
            }
    
    # Simple term
    return {
        'type': 'TERM',
        'term': query
    }


def apply_boolean_filter(df: pd.DataFrame, query: dict, column: str, exclusions: set = None) -> pd.DataFrame:
    """
    Apply a parsed Boolean query to filter a DataFrame.
    
    Args:
        df: DataFrame to filter
        query: Parsed Boolean query from parse_boolean_query()
        column: Column name to search in
        exclusions: Set of donor names to exclude from results
    
    Returns:
        Filtered DataFrame
    """
    if df.empty:
        return df
    
    if column not in df.columns:
        return df
    
    col_lower = df[column].str.lower().fillna('')
    
    if query['type'] == 'TERM':
        term = query['term'].lower()
        mask = col_lower.str.contains(term, na=False)
        result = df[mask].copy()
    
    elif query['type'] == 'OR':
        masks = []
        for sub_query in query['terms']:
            sub_result = apply_boolean_filter(df, sub_query, column)
            if not sub_result.empty:
                masks.append(df.index.isin(sub_result.index))
        
        if not masks:
            return pd.DataFrame()
        
        combined_mask = masks[0]
        for m in masks[1:]:
            combined_mask = combined_mask | m
        result = df[combined_mask].copy()
    
    elif query['type'] == 'NOT':
        term = query['term'].lower()
        mask = ~col_lower.str.contains(term, na=False)
        result = df[mask].copy()
    
    else:
        result = df
    
    # Apply exclusions if provided
    if exclusions and not result.empty and column in result.columns:
        result = result[~result[column].isin(exclusions)]
    
    return result


def get_search_terms(query: str) -> list:
    """Extract individual search terms from a Boolean query for API calls."""
    parsed = parse_boolean_query(query)
    terms = []
    
    def extract_terms(q):
        if q['type'] == 'TERM':
            terms.append(q['term'])
        elif q['type'] == 'OR':
            for sub in q['terms']:
                extract_terms(sub)
        elif q['type'] == 'NOT':
            terms.append(q['term'])
    
    extract_terms(parsed)
    return terms


def is_boolean_query(query: str) -> bool:
    """Check if query contains Boolean operators."""
    query_upper = query.upper()
    return ' OR ' in query_upper or query_upper.startswith('NOT ')


# ============= MAIN INTERFACE CONTINUED =============

# Initialize session state for exclusions and results
if 'excluded_donors' not in st.session_state:
    st.session_state.excluded_donors = set()
if 'raw_results' not in st.session_state:
    st.session_state.raw_results = {}
if 'last_query' not in st.session_state:
    st.session_state.last_query = ""

# Search input
col1, col2 = st.columns([3, 1])
with col1:
    search_query = st.text_input(
        "Enter company or donor name",
        placeholder="e.g., Google OR Microsoft OR Apple, NOT Anonymous",
        help="Search for donations. Supports: OR (any match), NOT (exclude)"
    )
    
    # Show Boolean query help
    if search_query and is_boolean_query(search_query):
        parsed = parse_boolean_query(search_query)
        if parsed['type'] == 'OR':
            st.caption(f"🔍 Boolean OR: Results contain ANY of the terms")
        elif parsed['type'] == 'NOT':
            st.caption(f"🔍 Boolean NOT: Excluding results containing '{parsed['term']}'")

with col2:
    search_button = st.button("🔍 Search", type="primary", use_container_width=True)

# Perform search - store raw results in session state
if search_button and search_query:
    # Clear exclusions for new search
    st.session_state.excluded_donors = set()
    st.session_state.last_query = search_query
    
    with st.spinner("Searching donation databases..."):
        raw_results = {}
        
        # Search all jurisdictions
        raw_results['uk'] = search_uk_donations(search_query, search_years)
        raw_results['germany'] = search_germany_donations(search_query, search_years)
        raw_results['austria'] = search_austria_donations(search_query)
        raw_results['italy'] = search_italy_donations(search_query)
        raw_results['netherlands'], raw_results['nl_full_data'] = search_netherlands_donations(search_query)
        raw_results['estonia'] = search_estonia_donations(search_query)
        raw_results['latvia'] = search_latvia_donations(search_query)
        raw_results['lithuania'] = search_lithuania_donations(search_query)
        raw_results['eu'] = search_eu_donations(search_query)
        
        st.session_state.raw_results = raw_results

# Display results if we have them
if st.session_state.raw_results and st.session_state.last_query:
    raw_results = st.session_state.raw_results
    
    # Collect all unique donor names for exclusion panel
    all_donors = set()
    donor_columns = {
        'uk': 'DonorName',
        'germany': 'Donor',
        'austria': 'Donor', 
        'italy': 'Donor',
        'netherlands': 'Donor',
        'estonia': 'Donor',
        'latvia': 'Donor',
        'lithuania': 'Donor',
        'eu': 'Donor'
    }
    
    for country, col in donor_columns.items():
        if country in raw_results and not raw_results[country].empty:
            df = raw_results[country]
            if col in df.columns:
                all_donors.update(df[col].dropna().unique())
    
    # Exclusion panel in sidebar
    if all_donors:
        with st.sidebar:
            st.write("### 🚫 Exclude False Positives")
            st.caption("Uncheck donors to remove from results & export")
            
            # Sort donors alphabetically
            sorted_donors = sorted(all_donors, key=str.lower)
            
            # Create checkboxes for each donor
            for donor in sorted_donors:
                is_excluded = donor in st.session_state.excluded_donors
                # Use checkbox - checked means INCLUDED
                include = st.checkbox(
                    donor[:50] + "..." if len(donor) > 50 else donor,
                    value=not is_excluded,
                    key=f"donor_{hash(donor)}"
                )
                if not include:
                    st.session_state.excluded_donors.add(donor)
                elif donor in st.session_state.excluded_donors:
                    st.session_state.excluded_donors.discard(donor)
            
            if st.session_state.excluded_donors:
                st.warning(f"Excluding {len(st.session_state.excluded_donors)} donor(s)")
                if st.button("Clear all exclusions"):
                    st.session_state.excluded_donors = set()
                    st.rerun()
    
    # Apply exclusions to get filtered results
    def apply_exclusions(df, donor_col):
        if df.empty or not st.session_state.excluded_donors:
            return df
        if donor_col in df.columns:
            return df[~df[donor_col].isin(st.session_state.excluded_donors)].copy()
        return df
    
    uk_results = apply_exclusions(raw_results.get('uk', pd.DataFrame()), 'DonorName')
    germany_results = apply_exclusions(raw_results.get('germany', pd.DataFrame()), 'Donor')
    austria_results = apply_exclusions(raw_results.get('austria', pd.DataFrame()), 'Donor')
    italy_results = apply_exclusions(raw_results.get('italy', pd.DataFrame()), 'Donor')
    netherlands_results = apply_exclusions(raw_results.get('netherlands', pd.DataFrame()), 'Donor')
    latvia_results = apply_exclusions(raw_results.get('latvia', pd.DataFrame()), 'Donor')
    estonia_results = apply_exclusions(raw_results.get('estonia', pd.DataFrame()), 'Donor')
    lithuania_results = apply_exclusions(raw_results.get('lithuania', pd.DataFrame()), 'Donor')
    eu_results = apply_exclusions(raw_results.get('eu', pd.DataFrame()), 'Donor')
    nl_full_data = raw_results.get('nl_full_data', False)
    
    # Show exclusion status
    if st.session_state.excluded_donors:
        st.info(f"🚫 **Excluding {len(st.session_state.excluded_donors)} donor(s):** {', '.join(sorted(st.session_state.excluded_donors))}")
    
    # Display UK results
    st.write("### 🇬🇧 United Kingdom")
    
    if not uk_results.empty:
        st.success(f"Found {len(uk_results)} donation records in UK")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = uk_results['ValueNumeric'].sum() if 'ValueNumeric' in uk_results.columns else 0
            st.metric("Total Value", f"£{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(uk_results))
        with col3:
            if 'RegulatedEntityName' in uk_results.columns:
                unique_parties = uk_results['RegulatedEntityName'].nunique()
                st.metric("Parties", unique_parties)
        
        formatted_uk = format_uk_results(uk_results)
        st.dataframe(formatted_uk, use_container_width=True, hide_index=True)
        
        if 'RegulatedEntityName' in uk_results.columns and 'ValueNumeric' in uk_results.columns:
            st.write("**Donations by Party:**")
            party_summary = uk_results.groupby('RegulatedEntityName').agg({
                'ValueNumeric': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (£)', 'Count']
            party_summary = party_summary.sort_values('Total (£)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No UK donation records found for this search term.")
    
    st.caption(f"**Note:** Showing last {search_years} years. Includes donations to parties and MPs.")
    
    st.divider()
    
    # Display Germany results
    st.write("### 🇩🇪 Germany")
    
    if not germany_results.empty:
        st.success(f"Found {len(germany_results)} donation records in Germany")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = germany_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(germany_results))
        with col3:
            unique_parties = germany_results['Party'].nunique()
            st.metric("Parties", unique_parties)
        
        formatted_de = format_germany_results(germany_results)
        st.dataframe(formatted_de, use_container_width=True, hide_index=True)
        
        if 'Party' in germany_results.columns:
            st.write("**Donations by Party:**")
            party_summary = germany_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No German donation records found for this search term.")
    
    st.caption("**Note:** German data covers large donations >€35,000 (since March 2024) or >€50,000 (before March 2024).")
    
    st.divider()
    
    # Display Austria results
    st.write("### 🇦🇹 Austria")
    
    if not austria_results.empty:
        st.success(f"Found {len(austria_results)} donation records in Austria")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = austria_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(austria_results))
        with col3:
            unique_parties = austria_results['Party'].nunique()
            st.metric("Parties", unique_parties)
        
        formatted_at = format_austria_results(austria_results)
        st.dataframe(formatted_at, use_container_width=True, hide_index=True)
        
        if 'Party' in austria_results.columns:
            st.write("**Donations by Party:**")
            party_summary = austria_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No Austrian donation records found for this search term.")
    
    st.caption("**Note:** Austrian data covers donations >€500. Threshold for immediate reporting: €2,500.")
    
    st.divider()
    
    # Display Italy results
    st.write("### 🇮🇹 Italy")
    
    if not italy_results.empty:
        st.success(f"Found {len(italy_results)} donation records in Italy")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = italy_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(italy_results))
        with col3:
            unique_parties = italy_results['Party'].nunique()
            st.metric("Parties", unique_parties)
        
        formatted_it = format_italy_results(italy_results)
        st.dataframe(formatted_it, use_container_width=True, hide_index=True)
        
        if 'Party' in italy_results.columns:
            st.write("**Donations by Party:**")
            party_summary = italy_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No Italian donation records found for this search term.")
    
    st.caption("**Note:** Italian data covers donations >€500. Data includes 2018-2024 from Parliament publications.")
    
    st.divider()
    
    # Display Netherlands results
    st.write("### 🇳🇱 Netherlands")
    
    # Show data source info
    if not nl_full_data:
        st.warning("⚠️ **Limited data mode:** Showing cached corporate/foundation donations only (47 key records). "
                  "For full data (789 records), install odfpy: `pip install odfpy`")
    
    if not netherlands_results.empty:
        st.success(f"Found {len(netherlands_results)} donation records in Netherlands")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = netherlands_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(netherlands_results))
        with col3:
            unique_parties = netherlands_results['Party'].nunique()
            st.metric("Parties", unique_parties)
        
        formatted_nl = format_netherlands_results(netherlands_results)
        st.dataframe(formatted_nl, use_container_width=True, hide_index=True)
        
        if 'Party' in netherlands_results.columns:
            st.write("**Donations by Party:**")
            party_summary = netherlands_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No Dutch donation records found for this search term.")
    
    st.caption("**Note:** Dutch data covers donations >€10,000 (2023-2025). Foreign donations are banned.")
    
    st.divider()
    
    # Display Latvia results
    st.write("### 🇱🇻 Latvia")
    
    if not latvia_results.empty:
        st.success(f"Found {len(latvia_results)} donation records in Latvia")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = latvia_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(latvia_results))
        with col3:
            unique_parties = latvia_results['Party'].nunique()
            st.metric("Parties", unique_parties)
        
        formatted_lv = format_latvia_results(latvia_results)
        st.dataframe(formatted_lv, use_container_width=True, hide_index=True)
        
        if 'Party' in latvia_results.columns:
            st.write("**Donations by Party:**")
            party_summary = latvia_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No Latvian donation records found for this search term.")
    
    st.caption("**Note:** Latvian data from KNAB (real-time). Corporate donations banned since 2004 - only natural persons can donate. [Full database →](https://info.knab.gov.lv/lv/db/ziedojumi/)")
    
    st.divider()
    
    # Display Estonia results
    st.write("### 🇪🇪 Estonia")
    
    if not estonia_results.empty:
        st.success(f"Found {len(estonia_results)} donation records in Estonia")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = estonia_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(estonia_results))
        with col3:
            unique_parties = estonia_results['Party'].nunique()
            st.metric("Parties", unique_parties)
        
        formatted_ee = format_estonia_results(estonia_results)
        st.dataframe(formatted_ee, use_container_width=True, hide_index=True)
        
        if 'Party' in estonia_results.columns:
            st.write("**Donations by Party:**")
            party_summary = estonia_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No Estonian donation records found for this search term.")
    
    st.caption("**Note:** Estonian data from ERJK (sample). Corporate donations banned - only natural persons can donate. [Full data →](https://www.erjk.ee/en)")
    
    st.divider()
    
    # Display Lithuania results
    st.write("### 🇱🇹 Lithuania")
    
    if not lithuania_results.empty:
        st.success(f"Found {len(lithuania_results)} donation records in Lithuania")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = lithuania_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(lithuania_results))
        with col3:
            unique_parties = lithuania_results['Party'].nunique()
            st.metric("Parties", unique_parties)
        
        formatted_lt = format_lithuania_results(lithuania_results)
        st.dataframe(formatted_lt, use_container_width=True, hide_index=True)
        
        if 'Party' in lithuania_results.columns:
            st.write("**Donations by Party:**")
            party_summary = lithuania_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No Lithuanian donation records found for this search term.")
    
    st.caption("**Note:** Lithuanian data from VRK (sample). Corporate donations banned - only citizens can donate. [Full data →](https://data.gov.lt/datasets/2016/)")
    
    st.divider()
    
    # Display EU results
    st.write("### 🇪🇺 European Union")
    
    if not eu_results.empty:
        st.success(f"Found {len(eu_results)} donation records at EU level")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            total_value = eu_results['Amount'].sum()
            st.metric("Total Value", f"€{total_value:,.2f}")
        with col2:
            st.metric("Number of Donations", len(eu_results))
        with col3:
            unique_parties = eu_results['Party'].nunique()
            st.metric("EU Parties", unique_parties)
        
        formatted_eu = format_eu_results(eu_results)
        st.dataframe(formatted_eu, use_container_width=True, hide_index=True)
        
        if 'Party' in eu_results.columns:
            st.write("**Donations by EU Party:**")
            party_summary = eu_results.groupby('Party').agg({
                'Amount': ['sum', 'count']
            }).round(2)
            party_summary.columns = ['Total (€)', 'Count']
            party_summary = party_summary.sort_values('Total (€)', ascending=False)
            st.dataframe(party_summary, use_container_width=True)
    else:
        st.info("No EU-level donation records found for this search term.")
    
    st.caption("**Note:** EU data covers donations to European political parties (e.g., EPP, PES, ALDE). Threshold: €12,000 for immediate disclosure.")
    
    # Add MEP resources info
    with st.expander("🔍 Looking for individual MEP data?"):
        st.markdown("""
        MEPs don't receive "donations" like national MPs - they declare **gifts** and **outside income** separately:
        
        - **[EU Integrity Watch](https://www.integritywatch.eu/mepincomes)** - MEP outside incomes & activities (Transparency International)
        - **[EP Gifts Register](https://www.europarl.europa.eu/meps/en/about/meps)** - Gifts >€150 received by MEPs
        - **[MEP Profile Pages](https://www.europarl.europa.eu/meps/en/home)** - Individual declarations of financial interests
        
        *MEPs must declare gifts >€150, outside earnings >€5,000/year, and meetings with lobbyists.*
        """)
    
    st.divider()
    
    # Export section
    st.write("### 📥 Export Results")
    
    # Show exclusion note if any
    if st.session_state.excluded_donors:
        st.info(f"📋 Export will exclude {len(st.session_state.excluded_donors)} donor(s): {', '.join(sorted(st.session_state.excluded_donors))}")
    
    # Generate Excel with filtered data
    excel_file = create_excel_report(st.session_state.last_query, uk_results, germany_results, austria_results, italy_results, netherlands_results, latvia_results, estonia_results, lithuania_results, eu_results)
    
    st.download_button(
        label="📊 Download Excel Report",
        data=excel_file,
        file_name=f"donations_{st.session_state.last_query.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
    
    st.caption("Excel report includes filtered data (exclusions applied), summary statistics, and source documentation.")

elif search_button:
    st.warning("Please enter a search term.")

# Footer
st.divider()
st.markdown("""
<div style='text-align: center; color: gray; font-size: 0.8em;'>
    Data sourced from official electoral commission databases. 
    For research and journalistic purposes.
</div>
""", unsafe_allow_html=True)
